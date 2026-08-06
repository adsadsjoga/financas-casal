"""
Gera o SQL da leva "CGD da Joana", a partir de
Joana_Ano_Completo_12M_Todos_Meses_Confirmados.xlsx.

Modo padrão — cria a conta CGD e importa as transações:

    python scripts/gerar_import_cgd_joana.py <xlsx> > supabase/aplicar/24_criar_conta_cgd_e_importar.sql

Sincronização de "Categoria atualizada" nas transações Revolut da Joana já
importadas (não mexe em CGD) — via CSV + tabela staging, não SQL monolítico
(colar ~600 KB de `values (...)` no SQL Editor do Supabase falhava com ERRO
42601, LINE 0 vazio — o navegador não entregava o paste inteiro):

    python scripts/gerar_import_cgd_joana.py --modo categorias-csv <xlsx> supabase/joana_categorias_excel.csv
    python scripts/gerar_import_cgd_joana.py --modo categorias-diagnostico > supabase/aplicar/26_diagnosticar_categorias_joana_passo1.sql
    python scripts/gerar_import_cgd_joana.py --modo categorias-update > supabase/aplicar/27_atualizar_categorias_joana_revolut.sql

Ordem de aplicação: `25_criar_staging_categorias.sql` (cria a tabela) ->
importar o CSV pelo Table Editor do Supabase (Insert > Import data from
CSV) -> `26_` (conferir os números) -> `27_` (aplicar) ->
`28_limpar_staging_categorias.sql` (apagar a tabela staging).

(`--modo categorias`, mantido como referência histórica, gera o SQL
monolítico antigo — não use para colar no SQL Editor, só documenta a
tentativa anterior.)

Rodar depois de conferir `supabase/aplicar/23_diagnosticar_joana_cgd.sql`
no SQL Editor — este script não acessa o banco, só lê o Excel.

DECISÕES QUE ESTE SCRIPT TOMA (e por quê):

* **Fonte única.** Só lê `Joana_Ano_Completo_12M_Todos_Meses_Confirmados.xlsx`
  — é estritamente mais completo que
  `Joana_Pagamentos_Nomeados_Classificados_e_Linkados.xlsx` (mesma aba
  "Joana Atualizada" com 96 linhas a mais) e é o único com a aba CGD.

* **Acesso às abas por índice, não por nome.** `wb.sheetnames` vem com
  encoding quebrado para os nomes acentuados deste arquivo específico
  ("Conciliação CGD-Revolut", "Parâmetros") — bug do exportador, não do
  openpyxl. Os índices abaixo foram conferidos manualmente contra o
  conteúdo real de cada aba.

* **Saldo inicial da conta CGD não é zero.** Ao contrário do Nubank (extrato
  sem saldo de abertura), a aba "Extratos CGD" traz 12 meses encadeados sem
  quebra, todos "Confirmado por PDF" — o saldo inicial real (187,38 € em
  2025-07-01) é dado confiável, não estimativa.

* **`external_id` = coluna `ID CGD`** (ex. `JCGD-202507-0003`), já única por
  linha na planilha — dedup exata, roda de novo sem duplicar.

* **Os 36 pares "Transferência interna" (CGD ↔ Revolut) entram como
  despesa/receita comuns, NÃO como `type='transferencia'`.** A aba
  "Conciliação CGD-Revolut" mostra que o lado Revolut de cada par já foi
  importado em 2026-08-03 como receita/despesa solta (não havia CGD no app
  ainda para virar transferência de verdade). Gravar agora o lado CGD como
  `type='transferencia'` apontando pro Revolut criaria um SEGUNDO movimento
  na conta Revolut para o mesmo dinheiro — dobraria o saldo dela. A entrada
  CGD vai para a categoria "Transferências internas" (nome mapeado por
  `CATEGORIA_MAP`, ver abaixo), que o dashboard já exclui do resultado.

* **`CATEGORIA_MAP`: "Transferência interna" (singular, nome usado só pelo
  CGD) vira "Transferências internas" (plural).** É o nome que
  `CATEGORIAS_FORA_DO_RESULTADO` (src/lib/constants.ts) reconhece pra
  excluir do dashboard. Importar com o nome singular recriaria o mesmo bug
  já corrigido em `08_unificar_categorias_duplicadas.sql` — duas categorias
  pro mesmo conceito, uma delas escapando do filtro por comparação literal.

* **Sincronização de categorias por FINGERPRINT calculado em SQL, não em
  Python.** O script não sabe o UUID das contas Revolut da Joana (não há
  acesso ao banco aqui) nem precisa saber: o SQL gerado resolve a conta por
  nome/dono e calcula o mesmo hash que o trigger `transactions_before_write()`
  grava, usando a função `public.normalize_description()` que já existe no
  banco — zero risco de o algoritmo divergir entre Python e Postgres.

* **Só atualiza automaticamente o que casa com exatamente 1 transação E tem
  `needs_review = true`.** Match ambíguo (fingerprint repetido — ex. duas
  compras idênticas no mesmo dia), sem match, ou transação já revisada
  manualmente (`needs_review = false`) com categoria diferente do Excel:
  fica só listado no SELECT de conferência, nunca em UPDATE automático.
"""

import csv
import re
import sys
import unicodedata

import openpyxl

# Índices confirmados manualmente (ver docstring acima) — não usar wb[nome]
# para as abas acentuadas.
ABA_CGD = 10               # "CGD Ano Completo"
ABA_EXTRATOS_CGD = 13       # "Extratos CGD"
ABA_JOANA_ATUALIZADA = 1    # "Joana Atualizada"

CATEGORIA_MAP = {
    "Transferência interna": "Transferências internas",
}

# Mapeamento (nome do Excel, tipo) -> categoria JÁ EXISTENTE no app, pra
# sincronização de categorias do Revolut da Joana. Levantado em 2026-08-06
# comparando as 75 combinações (categoria, tipo) que a aba "Joana
# Atualizada" usa contra a lista real de `public.categories` do casal —
# ver supabase/aplicar/README.md, seção "25b/26/27/28". Onde não há
# equivalente claro, cai em "Outras despesas"/"Outras receitas" (decisão
# do Gabriel: não criar categoria nova pra isso).
CATEGORIA_MAP_REVOLUT = {
    ("Ajuda Familiar", "despesa"): "Outras despesas",
    ("Ajuda Familiar", "receita"): "Outras receitas",
    # Faltou na primeira leva (2026-08-06): "Transferência interna" sem
    # sufixo (108 linhas, "To Joana Palminha" / "Payment from JOANA FILIPA
    # COSTA PALMINHA") caiu em "categoria não existe no app" no 26_ — o
    # mapeamento antigo (CATEGORIA_MAP, usado só pelo import do CGD) tinha
    # essa entrada, o novo dicionário específico do Revolut não.
    ("Transferência interna", "despesa"): "Transferências internas",
    ("Transferência interna", "receita"): "Transferências internas",
    ("Ajuste de Renda", "despesa"): "Outras despesas",
    ("Aporte para compra de veículo", "despesa"): "Carro",
    ("Assinaturas ", "despesa"): "Assinaturas",
    ("Assinaturas e telecomunicações", "despesa"): "Telefone e internet",
    ("Carro e Condução", "despesa"): "Carro",
    ("Carros - Gabriel", "despesa"): "Outras despesas",
    ("Chave de casa", "despesa"): "Moradia",
    ("Comida", "despesa"): "Alimentação",
    ("comida", "despesa"): "Alimentação",
    ("Compras Telemovel", "despesa"): "Compras",
    ("compras", "despesa"): "Compras",
    ("Compras e Roupa", "receita"): "Outras receitas",
    ("Cripto", "receita"): "Investimentos",
    ("Custo financeiro", "despesa"): "Empréstimos e Dívidas",
    ("Educacao", "despesa"): "Educação",
    ("Empréstimos e Dívidas", "receita"): "Outras receitas",
    ("Juros e Rendimentos Financeiros", "receita"): "Rendimentos",
    ("Juros/Dividendos", "receita"): "Rendimentos",
    ("lazer", "despesa"): "Lazer",
    ("Lazer e Entretenimento", "despesa"): "Lazer",
    ("Lazer e Entretenimento", "receita"): "Outras receitas",
    ("marketing", "despesa"): "Marketing",
    ("Moradia", "receita"): "Outras receitas",
    ("Negócio — Consoles/Vinted", "despesa"): "Outras despesas",
    ("Negócio — Consoles/Vinted", "receita"): "Vendas pessoais",
    ("Outros", "despesa"): "Outras despesas",
    ("Outros", "receita"): "Outras receitas",
    ("Presente", "despesa"): "Outras despesas",
    ("Renda/Entrada", "receita"): "Outras receitas",
    ("Restaurantes e Café", "despesa"): "Alimentação fora",
    ("Restaurantes e Café", "receita"): "Outras receitas",
    ("Revisar", "despesa"): "Outras despesas",
    ("Saúde e Farmácia", "despesa"): "Saúde",
    ("saude", "despesa"): "Saúde",
    ("Seguro", "despesa"): "Seguros",
    ("Seguro automóvel", "despesa"): "Seguros",
    ("Seguro automóvel", "receita"): "Outras receitas",
    ("Seguro pet", "despesa"): "Pets",
    ("Supermercado", "receita"): "Outras receitas",
    ("Trabalho", "despesa"): "Outras despesas",
    ("Transferência interna — Poupança", "despesa"): "Transferências internas",
    ("Transferência interna — Poupança", "receita"): "Transferências internas",
    ("Transferência interna — Trading 212", "despesa"): "Transferências internas",
    ("Transferência interna — Trading 212", "receita"): "Transferências internas",
    ("Transferência pessoal — revisar", "despesa"): "Outras despesas",
    ("Transferência pessoal — revisar", "receita"): "Outras receitas",
    ("Transporte", "receita"): "Outras receitas",
    ("Transporte Público", "despesa"): "Transporte",
    ("Transporte Público", "receita"): "Outras receitas",
    ("Viagem Casamento", "despesa"): "Viagem",
    ("Viagens", "despesa"): "Viagem",
    ("Viagens", "receita"): "Outras receitas",
    ("Viagens casamento", "despesa"): "Viagem",
    ("casamento", "despesa"): "Outras despesas",
    ("viagem casamento", "despesa"): "Viagem",
    ("viagem edimburgo", "despesa"): "Viagem",
    ("ginasio", "despesa"): "Ginasio",
}


def categoria_final_revolut(nome_excel: str, tipo: str) -> str:
    return CATEGORIA_MAP_REVOLUT.get((nome_excel, tipo), nome_excel)


# Mesma regra do Revolut: só categoria já existente no app, nunca cria nova.
# Levantado em 2026-08-06 comparando as 14 combinações (categoria, tipo) da
# aba "Joana Atualizada" filtrada por Banco='Trading 212'.
CATEGORIA_MAP_TRADING212 = {
    ("Cashback", "receita"): "Outras receitas",
    ("Compra de investimento", "despesa"): "Investimentos",
    ("Compras com cartão — revisar", "despesa"): "Outras despesas",
    ("Dividendos", "receita"): "Investimentos",
    ("Restaurantes e Café", "despesa"): "Alimentação fora",
    ("Resultado de investimentos", "receita"): "Investimentos",
    ("Taxas de investimento", "despesa"): "Investimentos",
    ("Taxas de investimento", "receita"): "Outras receitas",
    ("Transferência interna — Trading 212", "despesa"): "Transferências internas",
    ("Transferência interna — Trading 212", "receita"): "Transferências internas",
    ("Transporte Público", "despesa"): "Transporte",
    ("Viagens", "despesa"): "Viagem",
    # "Compras e Roupa" e "Supermercado" já batem direto com categoria existente.
}

# Categorias que, mesmo mapeadas, indicam que a própria planilha está
# incerta ("— revisar") — entram como needs_review=true independente do
# Estado da linha.
REVISAR_CATEGORIAS_TRADING212 = {"Compras com cartão — revisar"}

CATEGORIA_MAP_ACTIVOBANK = {
    ("A confirmar", "receita"): "Outras receitas",
    ("Compras e Eletrónica", "despesa"): "Compras",
    ("Impostos sobre investimentos", "despesa"): "Investimentos",
    ("Juros e Rendimentos Financeiros", "receita"): "Rendimentos",
    ("Negócio — Consoles/Vinted", "despesa"): "Outras despesas",
    ("Outros — revisar", "despesa"): "Outras despesas",
    ("Pagamento de empréstimo — Cofidis", "despesa"): "Empréstimos e Dívidas",
    ("Restaurantes e Café", "despesa"): "Alimentação fora",
    ("Saúde e Farmácia", "despesa"): "Saúde",
    ("Taxas bancárias", "despesa"): "Taxas bancárias",
    ("Transferência interna", "receita"): "Transferências internas",
    ("Transferência interna — Depósito a prazo", "despesa"): "Transferências internas",
    ("Transporte Público", "despesa"): "Transporte",
    # "Compras e Roupa" e "Supermercado" já batem direto com categoria existente.
}

REVISAR_CATEGORIAS_ACTIVOBANK = {"A confirmar", "Outros — revisar"}

# Estado da planilha que sinaliza incerteza mesmo com categoria definida —
# vira needs_review=true independente da categoria. "Reconciliado" e
# "Pareado" são os únicos estados "resolvidos"; qualquer outro vira revisão.
ESTADOS_RESOLVIDOS = {"Reconciliado", "Pareado"}


def gerar_import_banco(
    caminho: str,
    *,
    banco_excel: str,
    conta_nome: str,
    cor: str,
    categoria_map: dict,
    revisar_categorias: set,
) -> None:
    """Import genérico pra bancos da Joana sem saldo inicial confirmado por
    extrato (Trading 212, ActivoBank) — mesma estrutura do gerar_import_cgd(),
    mas sem saldo de referência pra conferir (só o fluxo líquido do período,
    que a aba 'Resumo 12M' também registra, usado como checagem cruzada)."""
    wb = openpyxl.load_workbook(caminho, data_only=True)
    idx, linhas = ler_aba(wb, ABA_JOANA_ATUALIZADA)

    registros = []
    sem_id = 0
    for r in linhas:
        if r[idx["Banco"]] != banco_excel:
            continue
        valor = r[idx["Valor"]]
        if not valor:
            continue
        tipo = "receita" if valor > 0 else "despesa"
        centavos = int(round(abs(valor) * 100))
        categoria_excel = r[idx["Categoria atualizada"]] or ""
        categoria = categoria_map.get((categoria_excel, tipo), categoria_excel)
        descricao = r[idx["Descrição"]] or ""
        data = r[idx["Data"]].date().isoformat()
        excel_id = r[idx["ID"]]
        if not excel_id:
            sem_id += 1
        estado = r[idx["Estado"]]
        revisar = (categoria_excel in revisar_categorias) or (estado not in ESTADOS_RESOLVIDOS)
        registros.append((data, tipo, centavos, descricao, excel_id, categoria, revisar))

    sys.stdout.reconfigure(encoding="utf-8")

    print(f"-- Importação {banco_excel} (Joana, EUR) — gerado por scripts/gerar_import_cgd_joana.py")
    print(f"-- {len(registros)} lançamentos.")
    print("--")
    print(f"-- Rodar INTEIRO de uma vez no SQL Editor do Supabase. Idempotente pra")
    print(f"-- {len(registros) - sem_id} linhas com 'ID' da planilha (external_id). As {sem_id} sem")
    print("-- ID (dividendos pequenos sem referência única) dependem só do fingerprint")
    print("-- automático do trigger — rodar este script duas vezes pode duplicá-las.")
    print("--")
    print("-- SEM saldo inicial confirmado por extrato (diferente do CGD) — a planilha")
    print("-- não traz reconciliação bancária pra esta conta. Saldo inicial = 0, mesma")
    print("-- decisão já tomada pro Nubank; ajustar depois manualmente em /contas.")
    print()
    print("begin;")
    print()

    print(f"-- 1. A conta. `accounts` não tem unique constraint em nome — `where not")
    print(f"--    exists` evita criar duas contas '{conta_nome}' se rodar duas vezes.")
    print(f"""insert into public.accounts
  (couple_id, name, type, currency, owner_profile_id, initial_balance_cents, color)
select c.id, '{conta_nome}', 'banco', 'EUR', m.profile_id, 0, '{cor}'
from public.couples c
join public.couple_members m on m.couple_id = c.id
join public.profiles p on p.id = m.profile_id
where p.display_name ilike 'joana%'
  and not exists (
    select 1 from public.accounts a where a.couple_id = c.id and a.name = '{conta_nome}'
  );""")
    print()

    categorias = sorted({(categoria, tipo) for (_, tipo, _, _, _, categoria, _) in registros})
    print("-- 2. Categorias — todas já deveriam existir no app (nenhuma nova por")
    print("--    desenho). `on conflict do nothing` só como rede de segurança:")
    for nome, tipo in categorias:
        print(f"--    - {nome} ({tipo})")
    print("insert into public.categories (couple_id, name, kind, icon)")
    print("select c.id, v.name, v.kind::public.category_kind, v.icon")
    print("from public.couples c, (values")
    print(",\n".join(f"  ('{escapar(n)}', '{t}', '{ICONES.get(n, '📦')}')" for n, t in categorias))
    print(") as v(name, kind, icon)")
    print("on conflict do nothing;")
    print()

    print("-- 3. Os lançamentos.")
    print(f"""with conta as (
       select a.id, a.couple_id from public.accounts a
       join public.profiles p on p.id = a.owner_profile_id
       where p.display_name ilike 'joana%' and a.name = '{conta_nome}' limit 1
     ),
     autor as (
       select m.profile_id from public.couple_members m
       join public.profiles p on p.id = m.profile_id
       where p.display_name ilike 'joana%' limit 1
     )
insert into public.transactions
  (couple_id, account_id, category_id, created_by, payer_profile_id, type,
   amount_cents, rate_to_primary, description, occurred_on, external_id, needs_review)
select conta.couple_id, conta.id,
       (select cat.id from public.categories cat
         where cat.couple_id = conta.couple_id
           and cat.name = v.categoria
           and cat.kind = v.tipo::public.category_kind
         limit 1),
       autor.profile_id, autor.profile_id, v.tipo::public.tx_type,
       v.centavos, 1, v.descricao, v.data::date, v.external_id, v.revisar
from conta, autor, (values""")

    valores = []
    revisar_n = 0
    for data, tipo, centavos, descricao, excel_id, categoria, revisar in registros:
        if revisar:
            revisar_n += 1
        external_id_sql = "null" if not excel_id else f"'{escapar(excel_id)}'"
        valores.append(
            f"  ('{data}', '{tipo}', {centavos}, '{escapar(descricao)}', "
            f"{external_id_sql}, '{escapar(categoria)}', {str(revisar).lower()})"
        )
    print(",\n".join(valores))
    print(") as v(data, tipo, centavos, descricao, external_id, categoria, revisar)")
    print("on conflict (account_id, external_id) where external_id is not null do nothing;")
    print()
    print("commit;")
    print()
    print("-- Conferência depois de rodar:")
    print(f"""-- select count(*) as lancamentos, count(*) filter (where needs_review) as revisar,
--        sum(case when t.type='receita' then t.amount_cents else -t.amount_cents end)/100.0
--          as variacao_liquida
-- from public.transactions t join public.accounts a on a.id = t.account_id
-- where a.name = '{conta_nome}';""")

    creditos = sum(c for _, tp, c, *_ in registros if tp == "receita")
    debitos = sum(c for _, tp, c, *_ in registros if tp == "despesa")

    print(f"-- {len(registros)} lançamentos; {revisar_n} marcados needs_review; "
          f"{sem_id} sem ID único na planilha.", file=sys.stderr)
    print(f"-- Créditos: {creditos / 100:.2f} EUR | Débitos: {debitos / 100:.2f} EUR "
          f"| Variação líquida: {(creditos - debitos) / 100:.2f} EUR", file=sys.stderr)
    print("-- Comparar contra a aba 'Resumo 12M' da planilha — não há saldo confirmado", file=sys.stderr)
    print("-- por PDF pra esta conta, só o fluxo do período.", file=sys.stderr)

ICONES = {
    "Transferências internas": "🔄",
    "Assinaturas e Digital": "📺",
    "Compras e Roupa": "🛍️",
    "Empréstimos e Dívidas": "💳",
    "Outros": "📦",
    "Pagamentos a pessoas": "🤝",
    "Recebimentos de pessoas": "🤝",
    "Supermercado": "🛒",
    "Taxas bancárias": "🏦",
}

MARCAS = re.compile(r"[̀-ͯ]")


def normalize_description(texto: str) -> str:
    """Espelha public.normalize_description() do schema (só usado para o log em stderr)."""
    sem_acento = MARCAS.sub("", unicodedata.normalize("NFD", texto or ""))
    minuscula = sem_acento.lower()
    colapsado = re.sub(r"[^a-z0-9]+", " ", minuscula)
    return re.sub(r"\s+", " ", colapsado).strip()


def escapar(texto) -> str:
    return str(texto).replace("'", "''")


def categoria_final(nome_excel: str) -> str:
    return CATEGORIA_MAP.get(nome_excel, nome_excel)


def ler_aba(wb, indice):
    ws = wb.worksheets[indice]
    linhas = list(ws.iter_rows(values_only=True))
    idx = {nome: i for i, nome in enumerate(linhas[0])}
    return idx, linhas[1:]


def gerar_import_cgd(caminho: str) -> None:
    wb = openpyxl.load_workbook(caminho, data_only=True)
    idx, linhas = ler_aba(wb, ABA_CGD)

    registros = []
    for r in linhas:
        net = r[idx["Movimento líquido (€)"]]
        if not net:
            continue
        tipo = "receita" if net > 0 else "despesa"
        centavos = int(round(abs(net) * 100))
        categoria_excel = r[idx["Categoria"]]
        categoria = categoria_final(categoria_excel)
        descricao = r[idx["Descrição original"]] or ""
        data = r[idx["Data valor"]].date().isoformat()
        excel_id = r[idx["ID CGD"]]
        revisar = categoria_excel == "Outros"
        registros.append((data, tipo, centavos, descricao, excel_id, categoria, revisar))

    idx_ext, linhas_ext = ler_aba(wb, ABA_EXTRATOS_CGD)
    saldo_inicial = float(linhas_ext[0][idx_ext["Saldo inicial (€)"]])
    saldo_final_esperado = float(linhas_ext[-1][idx_ext["Saldo final (€)"]])
    primeira_data = linhas_ext[0][idx_ext["Período inicial"]].date().isoformat()

    sys.stdout.reconfigure(encoding="utf-8")

    print("-- Importação CGD (Joana, EUR) — gerado por scripts/gerar_import_cgd_joana.py")
    print(f"-- {len(registros)} lançamentos, saldo inicial {saldo_inicial:.2f} € em {primeira_data}.")
    print("--")
    print("-- Rodar INTEIRO de uma vez no SQL Editor do Supabase, depois de conferir")
    print("-- 23_diagnosticar_joana_cgd.sql. Idempotente: a dedup usa o 'ID CGD' da")
    print("-- planilha (JCGD-AAAAMM-NNNN) como external_id.")
    print()
    print("begin;")
    print()

    print("-- 1. A conta. Saldo inicial vem do PDF do extrato (aba 'Extratos CGD'),")
    print(f"--    {saldo_inicial:.2f} € em {primeira_data} — os 12 meses da aba encadeiam")
    print("--    sem quebra e todos vêm 'Confirmado por PDF', não é estimativa.")
    print("--")
    print("--    `accounts` não tem unique constraint em nome — `where not exists` evita")
    print("--    criar duas contas 'CGD' se este script rodar duas vezes.")
    print(f"""insert into public.accounts
  (couple_id, name, type, currency, owner_profile_id, initial_balance_cents, color)
select c.id, 'CGD', 'banco', 'EUR', m.profile_id, {int(round(saldo_inicial * 100))}, '#06b6d4'
from public.couples c
join public.couple_members m on m.couple_id = c.id
join public.profiles p on p.id = m.profile_id
where p.display_name ilike 'joana%'
  and not exists (
    select 1 from public.accounts a where a.couple_id = c.id and a.name = 'CGD'
  );""")
    print()

    categorias = sorted({(categoria, tipo) for (_, tipo, _, _, _, categoria, _) in registros})
    print("-- 2. Categorias que o import usa, caso ainda não existam. CONFERIR contra o")
    print("--    resultado do passo 4 de 23_diagnosticar_joana_cgd.sql antes de rodar —")
    print("--    pode já existir uma parecida com nome diferente (mesmo risco que motivou")
    print("--    08_unificar_categorias_duplicadas.sql):")
    for nome, tipo in categorias:
        print(f"--    - {nome} ({tipo})")
    print("insert into public.categories (couple_id, name, kind, icon)")
    print("select c.id, v.name, v.kind::public.category_kind, v.icon")
    print("from public.couples c, (values")
    print(",\n".join(f"  ('{escapar(n)}', '{t}', '{ICONES.get(n, '📦')}')" for n, t in categorias))
    print(") as v(name, kind, icon)")
    print("on conflict do nothing;")
    print()

    print("-- 3. Os lançamentos.")
    print("""with conta as (
       select a.id, a.couple_id from public.accounts a
       join public.profiles p on p.id = a.owner_profile_id
       where p.display_name ilike 'joana%' and a.name = 'CGD' limit 1
     ),
     autor as (
       select m.profile_id from public.couple_members m
       join public.profiles p on p.id = m.profile_id
       where p.display_name ilike 'joana%' limit 1
     )
insert into public.transactions
  (couple_id, account_id, category_id, created_by, payer_profile_id, type,
   amount_cents, rate_to_primary, description, occurred_on, external_id, needs_review)
select conta.couple_id, conta.id,
       (select cat.id from public.categories cat
         where cat.couple_id = conta.couple_id
           and cat.name = v.categoria
           and cat.kind = v.tipo::public.category_kind
         limit 1),
       autor.profile_id, autor.profile_id, v.tipo::public.tx_type,
       v.centavos, 1, v.descricao, v.data::date, v.external_id, v.revisar
from conta, autor, (values""")

    valores = []
    revisar_n = 0
    for data, tipo, centavos, descricao, excel_id, categoria, revisar in registros:
        if revisar:
            revisar_n += 1
        valores.append(
            f"  ('{data}', '{tipo}', {centavos}, '{escapar(descricao)}', "
            f"'{escapar(excel_id)}', '{escapar(categoria)}', {str(revisar).lower()})"
        )
    print(",\n".join(valores))
    print(") as v(data, tipo, centavos, descricao, external_id, categoria, revisar)")
    print("on conflict (account_id, external_id) where external_id is not null do nothing;")
    print()
    print("commit;")
    print()
    print(f"-- Conferência depois de rodar (saldo deve fechar em {saldo_final_esperado:.2f} €):")
    print("""-- select a.name, count(*) as lancamentos,
--        a.initial_balance_cents / 100.0
--        + sum(case when t.type='receita' then t.amount_cents else -t.amount_cents end)/100.0
--          as saldo_final
-- from public.transactions t join public.accounts a on a.id = t.account_id
-- where a.name = 'CGD' group by a.name, a.initial_balance_cents;""")

    creditos = sum(c for _, tp, c, *_ in registros if tp == "receita")
    debitos = sum(c for _, tp, c, *_ in registros if tp == "despesa")
    saldo_calculado = saldo_inicial + (creditos - debitos) / 100

    print(f"-- {len(registros)} lançamentos; {revisar_n} marcados needs_review.", file=sys.stderr)
    print(f"-- Créditos: {creditos / 100:.2f} EUR | Débitos: {debitos / 100:.2f} EUR", file=sys.stderr)
    print(
        f"-- Saldo calculado: {saldo_calculado:.2f} EUR "
        f"(esperado, do extrato: {saldo_final_esperado:.2f} EUR)",
        file=sys.stderr,
    )
    if abs(saldo_calculado - saldo_final_esperado) > 0.01:
        print("-- ATENÇÃO: saldo calculado não bate com o extrato — NÃO rodar sem investigar.", file=sys.stderr)
    else:
        print("-- Saldo confere. Conferir mesmo assim contra o extrato real antes de rodar.", file=sys.stderr)


def montar_cte_excel(registros) -> str:
    valores = ",\n".join(
        f"  ({i}, '{data}'::date, '{tipo}', {centavos}, '{escapar(descricao)}', '{escapar(categoria)}')"
        for i, (data, tipo, centavos, descricao, categoria) in enumerate(registros)
    )
    # NÃO usa `fingerprint` (nem o hash equivalente calculado no Excel): as
    # transações da Joana no banco foram gravadas com um sufixo que a
    # planilha nova não tem, ex. banco = "Payment from DR JOHN CLARKE
    # [Categoria original Revolut: Top up]" vs Excel = "Payment from DR JOHN
    # CLARKE" — confirmado consultando o banco durante esta sessão. Qualquer
    # hash da descrição completa nunca bateria. Em vez disso casa por
    # data + valor + tipo + descrição normalizada como PREFIXO (a descrição
    # do Excel é sempre o início da descrição gravada, sufixo ou não).
    return f"""joana_revolut as (
  select a.id from public.accounts a
  join public.profiles p on p.id = a.owner_profile_id
  where p.display_name ilike 'joana%' and public.normalize_description(a.name) like '%revolut%'
),
excel(rowid, occurred_on, tipo, amount_cents, descricao, categoria) as (values
{valores}
),
-- LEFT JOIN de propósito: uma linha do Excel sem nenhuma transação
-- correspondente precisa continuar existindo aqui (com transaction_id nulo)
-- pra aparecer como "sem correspondência" no diagnóstico — com INNER JOIN
-- ela simplesmente desaparece do resultado em vez de ser contada.
candidatos_brutos as (
  select e.rowid, e.occurred_on, e.tipo, e.amount_cents, e.descricao, e.categoria,
         t.id as transaction_id, t.needs_review, cat_atual.name as categoria_atual
  from excel e
  cross join joana_revolut jr
  left join public.transactions t
    on t.account_id = jr.id
   and t.occurred_on = e.occurred_on
   and t.amount_cents = e.amount_cents
   and t.type = e.tipo::public.tx_type
   and public.normalize_description(t.description)
       like public.normalize_description(e.descricao) || '%'
  left join public.categories cat_atual on cat_atual.id = t.category_id
),
-- Quantas linhas do Excel (de qualquer rowid) apontaram pra essa mesma
-- transação real — se mais de uma, é ambíguo do lado do banco também.
match_por_transacao as (
  select transaction_id, count(*) as n
  from candidatos_brutos
  where transaction_id is not null
  group by 1
),
-- Por linha do Excel (rowid), soma quantas transações casaram — uma linha
-- do Excel não pode virar duas linhas de resultado só porque foi testada
-- contra as duas contas Revolut candidatas. count(transaction_id), não
-- count(*): ignora as linhas sem match (transaction_id nulo) na contagem.
por_linha as (
  select
    cb.rowid, cb.occurred_on, cb.tipo, cb.amount_cents, cb.descricao, cb.categoria,
    count(cb.transaction_id) as total_matches,
    min(cb.transaction_id::text)::uuid as transaction_id_unico,
    bool_or(coalesce(cb.needs_review, false)) as needs_review,
    max(cb.categoria_atual) as categoria_atual,
    max(mt.n) as max_linhas_por_transacao
  from candidatos_brutos cb
  left join match_por_transacao mt on mt.transaction_id = cb.transaction_id
  group by cb.rowid, cb.occurred_on, cb.tipo, cb.amount_cents, cb.descricao, cb.categoria
),
-- Só entra aqui quem casou com exatamente 1 transação, e essa transação só
-- foi alvo de 1 linha do Excel — evita pegar recorrência (duas compras
-- idênticas no mesmo dia) ou duplicata por engano.
candidatos as (
  select rowid, occurred_on, tipo, amount_cents, descricao, categoria,
         transaction_id_unico as transaction_id, needs_review, categoria_atual
  from por_linha
  where total_matches = 1 and max_linhas_por_transacao = 1
)"""


def ler_registros_categorias(caminho: str) -> tuple[list, int]:
    """(registros, ignoradas_sem_categoria) — linhas Revolut da Joana com
    categoria definida na aba 'Joana Atualizada'."""
    wb = openpyxl.load_workbook(caminho, data_only=True)
    idx, linhas = ler_aba(wb, ABA_JOANA_ATUALIZADA)

    registros = []
    ignoradas_sem_categoria = 0
    for r in linhas:
        if r[idx["Banco"]] != "Revolut":
            continue
        valor = r[idx["Valor"]]
        if not valor:
            continue
        categoria_excel = r[idx["Categoria atualizada"]]
        if not categoria_excel or categoria_excel == "A confirmar":
            ignoradas_sem_categoria += 1
            continue
        tipo = "receita" if valor > 0 else "despesa"
        centavos = int(round(abs(valor) * 100))
        descricao = r[idx["Descrição"]] or ""
        data = r[idx["Data"]].date().isoformat()
        registros.append((data, tipo, centavos, descricao, categoria_final_revolut(categoria_excel, tipo)))
    return registros, ignoradas_sem_categoria


def exportar_csv_categorias(caminho: str, saida_csv: str) -> None:
    """Gera o CSV pra importar em public._sync_joana_categorias pelo Table
    Editor do Supabase — substitui o `values (...)` gigante que falhava ao
    colar no SQL Editor (ERRO 42601, LINE 0 vazio: o paste de ~600 KB não
    chegava inteiro ao navegador)."""
    registros, ignoradas = ler_registros_categorias(caminho)
    with open(saida_csv, "w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(["rowid", "occurred_on", "tipo", "amount_cents", "descricao", "categoria"])
        for i, (data, tipo, centavos, descricao, categoria) in enumerate(registros):
            writer.writerow([i, data, tipo, centavos, descricao, categoria])
    print(f"-- {len(registros)} linhas escritas em {saida_csv} "
          f"({ignoradas} ficaram de fora por estar 'A confirmar' ou vazias).", file=sys.stderr)
    print("-- Próximo passo: rodar 25_criar_staging_categorias.sql, depois importar", file=sys.stderr)
    print(f"-- {saida_csv} pelo Table Editor do Supabase (Insert > Import data from CSV).", file=sys.stderr)


def montar_cte_staging() -> str:
    """Mesma lógica de casamento de montar_cte_excel(), mas lendo os dados
    de public._sync_joana_categorias em vez de embutir um `values (...)`
    com milhares de linhas — SQL final fica curto o bastante pra colar sem
    risco de o paste falhar no navegador."""
    return """joana_revolut as (
  select a.id, a.couple_id from public.accounts a
  join public.profiles p on p.id = a.owner_profile_id
  where p.display_name ilike 'joana%' and public.normalize_description(a.name) like '%revolut%'
),
excel as (
  select rowid, occurred_on, tipo, amount_cents, descricao, categoria
  from public._sync_joana_categorias
),
-- LEFT JOIN de propósito: uma linha do Excel sem nenhuma transação
-- correspondente precisa continuar existindo aqui (com transaction_id nulo)
-- pra aparecer como "sem correspondência" no diagnóstico — com INNER JOIN
-- ela simplesmente desaparece do resultado em vez de ser contada. Mesma
-- coisa pro `cat_alvo`: categoria do Excel que não existe em
-- `public.categories` tem que aparecer como problema explícito no
-- diagnóstico, não sumir do UPDATE sem avisar (foi exatamente isso que
-- aconteceu na primeira tentativa, 2026-08-06 — o UPDATE com INNER JOIN
-- em categories ignorava silenciosamente quem não tinha categoria, e o
-- segundo UPDATE zerava needs_review de qualquer jeito).
candidatos_brutos as (
  select e.rowid, e.occurred_on, e.tipo, e.amount_cents, e.descricao, e.categoria,
         t.id as transaction_id, t.needs_review, cat_atual.name as categoria_atual,
         cat_alvo.id as categoria_alvo_id
  from excel e
  cross join joana_revolut jr
  left join public.transactions t
    on t.account_id = jr.id
   and t.occurred_on = e.occurred_on
   and t.amount_cents = e.amount_cents
   and t.type = e.tipo::public.tx_type
   and public.normalize_description(t.description)
       like public.normalize_description(e.descricao) || '%'
  left join public.categories cat_atual on cat_atual.id = t.category_id
  left join public.categories cat_alvo
    on cat_alvo.couple_id = jr.couple_id
   and cat_alvo.name = e.categoria
   and cat_alvo.kind = e.tipo::public.category_kind
),
-- Quantas linhas do Excel (de qualquer rowid) apontaram pra essa mesma
-- transação real — se mais de uma, é ambíguo do lado do banco também.
match_por_transacao as (
  select transaction_id, count(*) as n
  from candidatos_brutos
  where transaction_id is not null
  group by 1
),
-- Por linha do Excel (rowid), soma quantas transações casaram — uma linha
-- do Excel não pode virar duas linhas de resultado só porque foi testada
-- contra as duas contas Revolut candidatas. count(transaction_id), não
-- count(*): ignora as linhas sem match (transaction_id nulo) na contagem.
por_linha as (
  select
    cb.rowid, cb.occurred_on, cb.tipo, cb.amount_cents, cb.descricao, cb.categoria,
    count(cb.transaction_id) as total_matches,
    min(cb.transaction_id::text)::uuid as transaction_id_unico,
    bool_or(coalesce(cb.needs_review, false)) as needs_review,
    max(cb.categoria_atual) as categoria_atual,
    max(mt.n) as max_linhas_por_transacao,
    bool_or(cb.categoria_alvo_id is not null) as categoria_alvo_existe
  from candidatos_brutos cb
  left join match_por_transacao mt on mt.transaction_id = cb.transaction_id
  group by cb.rowid, cb.occurred_on, cb.tipo, cb.amount_cents, cb.descricao, cb.categoria
),
-- Só entra aqui quem casou com exatamente 1 transação, essa transação só
-- foi alvo de 1 linha do Excel (evita recorrência/duplicata), E a
-- categoria alvo existe em public.categories — sem essa última checagem
-- o UPDATE do 27_ (que faz INNER JOIN em categories) ignoraria a linha
-- sem avisar, exatamente o bug encontrado em 2026-08-06.
candidatos as (
  select rowid, occurred_on, tipo, amount_cents, descricao, categoria,
         transaction_id_unico as transaction_id, needs_review, categoria_atual
  from por_linha
  where total_matches = 1 and max_linhas_por_transacao = 1 and categoria_alvo_existe
)"""


def gerar_diagnostico_staging() -> None:
    """26_diagnosticar_categorias_joana_passo1.sql — lê de public._sync_joana_categorias
    (importado do CSV), não embute dados. Rodar depois de 25_ + import do CSV."""
    sys.stdout.reconfigure(encoding="utf-8")
    cte = montar_cte_staging()

    print("-- PASSO 1 — diagnóstico da sincronização de categorias (Revolut da Joana)")
    print("-- Lê de public._sync_joana_categorias — rodar depois de 25_criar_staging_categorias.sql")
    print("-- e de importar joana_categorias_excel.csv pelo Table Editor.")
    print("--")
    print("-- Casa cada linha por data + valor + tipo + descrição normalizada como PREFIXO")
    print("-- (não por fingerprint: a descrição gravada no banco tem um sufixo")
    print("-- '[Categoria original Revolut: ...]' que a planilha não tem).")
    print("--")
    print("-- Se alguma linha vier como 'categoria não existe no app' o UPDATE do 27_")
    print("-- NÃO vai tocar nela (INNER JOIN em categories) — se isso acontecer, é sinal")
    print("-- de que CATEGORIA_MAP_REVOLUT no script Python precisa de mais uma entrada.")
    print("-- Se vier diferente do que já foi conferido antes, o CSV pode ter subido")
    print("-- errado (linha faltando, delimitador, encoding) — parar antes do 27_.")
    print()
    print(f"with {cte}")
    print("select")
    print("  case")
    print("    when total_matches = 0 then 'sem correspondência'")
    print("    when total_matches > 1 or max_linhas_por_transacao > 1")
    print("      then 'correspondência ambígua (recorrência ou duplicata)'")
    print("    when not categoria_alvo_existe then 'categoria não existe no app'")
    print("    when not needs_review and categoria_atual is distinct from categoria")
    print("      then 'será atualizada (já revisada, categoria diverge)'")
    print("    when categoria_atual is not distinct from categoria then 'já está correto'")
    print("    else 'será atualizada'")
    print("  end as situacao,")
    print("  count(*) as linhas,")
    print("  (array_agg(descricao order by occurred_on))[1:5] as exemplos")
    print("from por_linha")
    print("group by 1")
    print("order by 1;")


def gerar_update_staging() -> None:
    """27_atualizar_categorias_joana_revolut.sql — lê de public._sync_joana_categorias.
    Rodar só depois de conferir os números do 26_."""
    sys.stdout.reconfigure(encoding="utf-8")
    cte = montar_cte_staging()

    print("-- PASSO 2 — aplica a sincronização de categorias. Rodar só depois de")
    print("-- conferir 26_diagnosticar_categorias_joana_passo1.sql.")
    print("--")
    print("-- Só toca em transação com match único (nunca ambíguo, nunca sem")
    print("-- correspondência). Inclui needs_review=false por decisão explícita do")
    print("-- Gabriel em 2026-08-06 — o balde 'já revisada, categoria diverge' tinha")
    print("-- 1.422 linhas e ele preferiu aplicar tudo de uma vez a deixar pendente.")
    print()
    print("begin;")
    print()
    print(f"with {cte}")
    print("update public.transactions t")
    print("set category_id = novo.category_id")
    print("from (")
    print("  select c.transaction_id, cat.id as category_id")
    print("  from candidatos c")
    print("  join public.transactions tx on tx.id = c.transaction_id")
    print("  join public.categories cat")
    print("    on cat.couple_id = tx.couple_id")
    print("   and cat.name = c.categoria")
    print("   and cat.kind = c.tipo::public.category_kind")
    print("  where c.categoria_atual is distinct from c.categoria")
    print(") novo")
    print("where t.id = novo.transaction_id;")
    print()
    print(f"with {cte}")
    print("update public.transactions t set needs_review = false")
    print("from candidatos c")
    print("where t.id = c.transaction_id and c.needs_review;")
    print()
    print("commit;")


def gerar_sync_categorias(caminho: str) -> None:
    """Versão original — SQL monolítico com os dados embutidos em `values (...)`.
    Mantida só como referência histórica (ver supabase/aplicar/README.md,
    seção 25.): colar ~600 KB de SQL no SQL Editor do Supabase falhava com
    ERRO 42601 (LINE 0 vazio). Superada por exportar_csv_categorias() +
    gerar_diagnostico_staging() + gerar_update_staging()."""
    registros, ignoradas_sem_categoria = ler_registros_categorias(caminho)
    sys.stdout.reconfigure(encoding="utf-8")
    cte = montar_cte_excel(registros)

    print("-- Sincronização de categorias — Revolut da Joana")
    print("-- gerado por scripts/gerar_import_cgd_joana.py --modo categorias")
    print(f"-- {len(registros)} linhas do Excel com categoria definida "
          f"({ignoradas_sem_categoria} ficaram de fora por estar 'A confirmar' ou vazias).")
    print("--")
    print("-- Casa cada linha do Excel com uma transação já importada por")
    print("-- data + valor + tipo + descrição normalizada como PREFIXO (não por")
    print("-- fingerprint: a descrição gravada no banco tem um sufixo")
    print("-- '[Categoria original Revolut: ...]' que a planilha não tem — conferido")
    print("-- em produção nesta sessão, 2026-08-06). Calculado em SQL puro contra as")
    print("-- duas contas Revolut da Joana — não precisa do UUID delas de antemão.")
    print("--")
    print("-- ===========================================================================")
    print("-- PASSO 1 — rode só o SELECT abaixo primeiro e leia o resultado. Só as linhas")
    print("-- 'sem correspondência' e 'correspondência ambígua' NÃO são tocadas pelo")
    print("-- UPDATE do passo 2. As demais categorias diferentes da atual SÃO")
    print("-- atualizadas, inclusive as já revisadas manualmente — decisão explícita do")
    print("-- Gabriel em 2026-08-06 (o padrão original só tocava needs_review=true; ele")
    print("-- preferiu aplicar tudo de uma vez depois de ver que o balde 'já revisada'")
    print("-- tinha 1.422 linhas).")
    print("-- ===========================================================================")
    print()
    print(f"with {cte}")
    print("select")
    print("  case")
    print("    when total_matches = 0 then 'sem correspondência'")
    print("    when total_matches > 1 or max_linhas_por_transacao > 1")
    print("      then 'correspondência ambígua (recorrência ou duplicata)'")
    print("    when not needs_review and categoria_atual is distinct from categoria")
    print("      then 'será atualizada (já revisada, categoria diverge)'")
    print("    when categoria_atual is not distinct from categoria then 'já está correto'")
    print("    else 'será atualizada'")
    print("  end as situacao,")
    print("  count(*) as linhas,")
    print("  (array_agg(descricao order by occurred_on))[1:5] as exemplos")
    print("from por_linha")
    print("group by 1")
    print("order by 1;")
    print()
    print("-- ===========================================================================")
    print("-- PASSO 2 — só depois de conferir o SELECT acima, rode o bloco abaixo inteiro.")
    print("-- Só toca em transação com match único (nunca ambíguo, nunca sem")
    print("-- correspondência). Inclui needs_review=false por decisão explícita do")
    print("-- Gabriel em 2026-08-06 — o SELECT do PASSO 1 mostrou 1.422 linhas nesse")
    print("-- balde e ele preferiu aplicar tudo de uma vez a deixar pendente.")
    print("-- ===========================================================================")
    print()
    print("begin;")
    print()
    print(f"with {cte}")
    print("update public.transactions t")
    print("set category_id = novo.category_id")
    print("from (")
    print("  select c.transaction_id, cat.id as category_id")
    print("  from candidatos c")
    print("  join public.transactions tx on tx.id = c.transaction_id")
    print("  join public.categories cat")
    print("    on cat.couple_id = tx.couple_id")
    print("   and cat.name = c.categoria")
    print("   and cat.kind = c.tipo::public.category_kind")
    print("  where c.categoria_atual is distinct from c.categoria")
    print(") novo")
    print("where t.id = novo.transaction_id;")
    print()
    print(f"with {cte}")
    print("update public.transactions t set needs_review = false")
    print("from candidatos c")
    print("where t.id = c.transaction_id and c.needs_review;")
    print()
    print("commit;")

    print(f"-- {len(registros)} linhas candidatas no Excel; "
          f"{ignoradas_sem_categoria} sem categoria definida (ignoradas).", file=sys.stderr)
    print("-- Rode o PASSO 1 (SELECT) primeiro e confira a contagem de "
          "'sem correspondência'/'revisar manualmente' antes do PASSO 2.", file=sys.stderr)


def main() -> None:
    args = sys.argv[1:]
    modo = "import"
    if args and args[0] == "--modo":
        modo = args[1]
        args = args[2:]

    if modo == "import":
        if not args:
            sys.exit("uso: python scripts/gerar_import_cgd_joana.py <xlsx>")
        gerar_import_cgd(args[0])
    elif modo == "categorias":
        if not args:
            sys.exit("uso: python scripts/gerar_import_cgd_joana.py --modo categorias <xlsx>")
        gerar_sync_categorias(args[0])
    elif modo == "categorias-csv":
        if len(args) < 2:
            sys.exit(
                "uso: python scripts/gerar_import_cgd_joana.py --modo categorias-csv <xlsx> <saida.csv>"
            )
        exportar_csv_categorias(args[0], args[1])
    elif modo == "categorias-diagnostico":
        gerar_diagnostico_staging()
    elif modo == "categorias-update":
        gerar_update_staging()
    elif modo == "trading212":
        if not args:
            sys.exit("uso: python scripts/gerar_import_cgd_joana.py --modo trading212 <xlsx>")
        gerar_import_banco(
            args[0],
            banco_excel="Trading 212",
            conta_nome="Trading 212",
            cor="#f59e0b",
            categoria_map=CATEGORIA_MAP_TRADING212,
            revisar_categorias=REVISAR_CATEGORIAS_TRADING212,
        )
    elif modo == "activobank":
        if not args:
            sys.exit("uso: python scripts/gerar_import_cgd_joana.py --modo activobank <xlsx>")
        gerar_import_banco(
            args[0],
            banco_excel="ActivoBank",
            conta_nome="ActivoBank",
            cor="#14b8a6",
            categoria_map=CATEGORIA_MAP_ACTIVOBANK,
            revisar_categorias=REVISAR_CATEGORIAS_ACTIVOBANK,
        )
    else:
        sys.exit(
            f"modo desconhecido: {modo!r} (use 'import', 'categorias', 'categorias-csv', "
            "'categorias-diagnostico', 'categorias-update', 'trading212' ou 'activobank')"
        )


if __name__ == "__main__":
    main()
